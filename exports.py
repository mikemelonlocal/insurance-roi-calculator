"""Consolidated export builders for HTML, PDF, and Excel."""

import io

import pandas as pd
import streamlit as st

from utils import fmt, now_local, get_tz_label
from config import DARK_GREEN, PRI_GREEN
from charts import plotly_to_matplotlib_image

# Optional PDF support
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ── HTML helpers ─────────────────────────────────────────────────────────────

_HTML_CSS = (
    'body{font-family:Poppins,sans-serif;background:#FFFFFF;color:#1a1a1a;margin:0;padding:32px;}'
    '.topbar{background:#114E38;color:#fff;padding:18px 32px;margin:-32px -32px 32px -32px;'
    'font-size:1.3rem;font-weight:700;}'
    'h1{color:#114E38;font-size:1.6rem;font-weight:700;margin-bottom:4px;}'
    'h2{color:#114E38;font-size:1.1rem;font-weight:600;margin-top:28px;margin-bottom:8px;'
    'border-bottom:2px solid #114E38;padding-bottom:4px;}'
    '.subtitle{color:#666;font-size:0.85rem;margin-bottom:20px;}'
    'table{border-collapse:collapse;width:100%;margin-top:12px;}'
    'th{background:#114E38;color:#fff;padding:10px 14px;text-align:left;font-size:0.82rem;font-weight:600;}'
    'td{padding:9px 14px;border-bottom:1px solid #D9D9D9;font-size:0.88rem;}'
    'tr:nth-child(even) td{background:#FEF8E9;}'
    'tr.total td{font-weight:700;background:#EDDFDB;}'
    '.metric-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:12px;margin-top:12px;}'
    '.metric-card{background:#FFFFFF;border-radius:8px;padding:12px 16px;border:1px solid #D9D9D9;}'
    '.metric-label{font-size:0.75rem;color:#114E38;font-weight:600;margin-bottom:4px;}'
    '.metric-value{font-size:1.3rem;font-weight:700;color:#1a1a1a;}'
    '.prod-card{background:#FEF8E9;border-left:4px solid #114E38;border-radius:0 8px 8px 0;padding:14px 18px;margin-bottom:12px;}'
    '.prod-title{font-size:1rem;font-weight:700;color:#114E38;margin-bottom:8px;}'
    '.total-box{border-radius:8px;padding:16px 20px;margin-top:16px;color:#fff;}'
    '.total-box .label{font-size:0.85rem;opacity:0.9;font-weight:600;}'
    '.total-box .value{font-size:1.7rem;font-weight:700;}'
    '.total-box .sub{font-size:0.78rem;opacity:0.85;margin-top:4px;}'
    '.two-col{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-top:12px;}'
    '.winner{background:#FEF8E9;border-left:4px solid #47B74F;border-radius:0 8px 8px 0;'
    'padding:12px 16px;color:#114E38;font-weight:600;margin-top:12px;}'
    '.callout{background:#FEF8E9;border-radius:8px;padding:14px 18px;margin-bottom:12px;}'
    '.footer{margin-top:40px;font-size:0.75rem;color:#999;border-top:1px solid #D9D9D9;padding-top:12px;}'
    '.tag-tight{background:#c8e6c9;padding:2px 8px;border-radius:4px;font-size:0.8rem;}'
    '.tag-mod{background:#fff9c4;padding:2px 8px;border-radius:4px;font-size:0.8rem;}'
    '.tag-over{background:#ffe0b2;padding:2px 8px;border-radius:4px;font-size:0.8rem;}'
    '@media (max-width: 768px) {'
    '.two-col{grid-template-columns:1fr;}'
    '.metric-grid{grid-template-columns:1fr;}'
    '}'
)


def html_wrap(title, body):
    """Wrap *body* HTML in a branded template."""
    ts = now_local().strftime(f"%Y-%m-%d %I:%M %p {get_tz_label()}")
    return (
        '<!DOCTYPE html><html><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
        '<link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;700&display=swap" rel="stylesheet">'
        f'<title>{title}</title>'
        f'<style>{_HTML_CSS}</style></head><body>'
        '<div class="topbar">Insurance ROI Calculator</div>'
        + body +
        f'<div class="footer">Generated {ts} by Insurance ROI Calculator — Melon Local</div>'
        '</body></html>'
    )


def metric_card_html(label, value):
    return (
        '<div class="metric-card">'
        f'<div class="metric-label">{label}</div>'
        f'<div class="metric-value">{value}</div>'
        '</div>'
    )


def build_html_table(headers, rows, total_row=None):
    """Build an HTML <table> string from headers and rows."""
    html = '<table><thead><tr>'
    for h in headers:
        html += f'<th>{h}</th>'
    html += '</tr></thead><tbody>'
    for row in rows:
        html += '<tr>'
        for cell in row:
            html += f'<td>{cell}</td>'
        html += '</tr>'
    if total_row:
        html += '<tr class="total">'
        for cell in total_row:
            html += f'<td><strong>{cell}</strong></td>'
        html += '</tr>'
    html += '</tbody></table>'
    return html


# ── PDF builder ──────────────────────────────────────────────────────────────

def build_pdf_report(title, subtitle, table_data, col_widths, chart_fig=None):
    """Build a PDF report. Returns bytes or None if reportlab is unavailable."""
    if not REPORTLAB_AVAILABLE:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                 fontSize=24, textColor=colors.HexColor(DARK_GREEN),
                                 spaceAfter=12, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'],
                                    fontSize=12, textColor=colors.HexColor(PRI_GREEN),
                                    spaceAfter=20, alignment=TA_CENTER)

    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(DARK_GREEN)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF8E9')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
    ]))
    elements.append(table)

    if chart_fig is not None:
        try:
            chart_buf = plotly_to_matplotlib_image(chart_fig, width_inches=6.5, height_inches=3.5)
            if chart_buf:
                elements.append(Spacer(1, 0.3 * inch))
                chart_img = RLImage(chart_buf, width=6.5 * inch, height=3.5 * inch)
                elements.append(chart_img)
        except Exception:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph("(Chart not available — see HTML export)", styles['Italic']))

    elements.append(Spacer(1, 0.3 * inch))
    ts = now_local().strftime(f'%B %d, %Y at %I:%M %p {get_tz_label()}')
    elements.append(Paragraph(f"Generated: {ts}", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ── Excel builder ────────────────────────────────────────────────────────────

def create_formatted_excel(df, sheet_name='Data'):
    """Create a branded Excel file from a DataFrame. Returns bytes or None."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        return None

    try:
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3'),
        )

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 1:
                    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='114E38', end_color='114E38', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='Calibri', size=11)
                    cell.alignment = Alignment(
                        horizontal='left' if c_idx == 1 else 'right',
                        vertical='center',
                    )

                    if r_idx > 1:
                        header = ws.cell(row=1, column=c_idx).value
                        if header and isinstance(value, (int, float)):
                            hdr = str(header).lower()
                            if '%' in str(header) or 'rate' in hdr:
                                cell.number_format = '0.0%' if value < 1 else '0.0"%"'
                            elif any(w in hdr for w in ['cost', 'price', 'revenue', 'premium', 'payroll', 'total', 'commission']):
                                cell.number_format = '$#,##0.00'

        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        return None
